"""INEIDataSource — precios de referencia de los Índices Unificados (ADR-002).

Lee los precios desde la tabla `precios_referencia` (alimentada por el ETL).
Incluye un cargador de .xlsx para el pipeline de ingesta mensual.
"""

from datetime import datetime
from decimal import Decimal
from typing import Optional

from app.core.database import SessionLocal
from app.models.precio_referencia import PrecioReferencia


class INEIDataSource:
    """Fuente de precios de referencia INEI, con resolución regional → nacional."""

    def obtener_precio_referencia(
        self, codigo_inei: str, departamento: Optional[str] = None
    ) -> Optional[PrecioReferencia]:
        db = SessionLocal()
        try:
            if departamento:
                regional = (
                    db.query(PrecioReferencia)
                    .filter(
                        PrecioReferencia.codigo_inei == codigo_inei,
                        PrecioReferencia.departamento == departamento,
                        PrecioReferencia.fuente == "inei",
                    )
                    .order_by(PrecioReferencia.anio.desc(), PrecioReferencia.mes.desc())
                    .first()
                )
                if regional:
                    return regional
            return (
                db.query(PrecioReferencia)
                .filter(
                    PrecioReferencia.codigo_inei == codigo_inei,
                    PrecioReferencia.departamento.is_(None),
                    PrecioReferencia.fuente == "inei",
                )
                .order_by(PrecioReferencia.anio.desc(), PrecioReferencia.mes.desc())
                .first()
            )
        finally:
            db.close()

    def obtener_costo_m2(self, tipo_obra: str, departamento: str) -> Optional[Decimal]:
        # El costo/m² lo provee MVCS; INEI solo precios por insumo.
        return None

    def _insertar_en_db(self, registros: list[PrecioReferencia]) -> int:
        """Inserta registros de precios en BD. Separado para facilitar tests mock."""
        db = SessionLocal()
        try:
            for r in registros:
                db.add(r)
            db.commit()
            return len(registros)
        finally:
            db.close()

    def cargar_xlsx(self, ruta_xlsx: str, mes: int, anio: int) -> int:
        """Parsea un .xlsx de Índices Unificados y hace upsert en BD.

        Espera columnas: codigo_inei, insumo, unidad, precio, departamento.
        Retorna el número de filas cargadas.
        """
        from openpyxl import load_workbook

        wb = load_workbook(ruta_xlsx, read_only=True, data_only=True)
        hoja = wb.active
        filas = list(hoja.iter_rows(min_row=2, values_only=True))

        registros: list[PrecioReferencia] = []
        for codigo, insumo, unidad, precio, departamento in filas:
            if not codigo or precio is None:
                continue
            registros.append(
                PrecioReferencia(
                    codigo_inei=str(codigo),
                    insumo=str(insumo or ""),
                    unidad=str(unidad or ""),
                    precio=Decimal(str(precio)),
                    departamento=departamento or None,
                    mes=mes,
                    anio=anio,
                    fuente="inei",
                )
            )
        if registros:
            return self._insertar_en_db(registros)
        return 0

    def get_metadata(self) -> dict:
        return {"source": "inei", "tipo": "indices_unificados"}

    def get_last_update(self) -> Optional[datetime]:
        db = SessionLocal()
        try:
            ultimo = (
                db.query(PrecioReferencia)
                .filter(PrecioReferencia.fuente == "inei")
                .order_by(PrecioReferencia.anio.desc(), PrecioReferencia.mes.desc())
                .first()
            )
            if ultimo and ultimo.anio and ultimo.mes:
                return datetime(ultimo.anio, ultimo.mes, 1)
            return None
        finally:
            db.close()
